from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter
from doc2md.processors.base import BaseProcessor
from doc2md.models import ConversionResult
from doc2md.config import config

class XlsxProcessor(BaseProcessor):
    SUPPORTED_EXTENSIONS = [".xlsx", ".xls", ".csv"]

    def process(self, file_path: str) -> ConversionResult:
        filename = Path(file_path).stem
        self.assets = []

        # Extract images from xlsx (charts, etc.)
        self._extract_images(file_path, filename)

        # Determine format and process
        ext = Path(file_path).suffix.lower()
        if ext == ".csv":
            markdown = self._process_csv(file_path)
        else:
            markdown = self._process_xlsx(file_path)

        # Save markdown
        output_dir = Path(config.PROJECTS_DIR) / self.project_id
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / f"{filename}.md"
        output_path.write_text(markdown)

        return ConversionResult(
            markdown=markdown,
            assets=self.assets,
            project_id=self.project_id,
            filename=f"{filename}.md"
        )

    def _process_xlsx(self, file_path: str) -> str:
        """Process Excel file to Markdown."""
        wb = openpyxl.load_workbook(file_path)
        sheets = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            lines = [f"## {sheet_name}\n"]

            for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                cells = [str(cell) if cell is not None else "" for cell in row]
                lines.append(f"| {' | '.join(cells) } |")

                if row_idx == 1:
                    lines.append("|" + "|".join(["---"] * len(cells)) + "|")

            sheets.append("\n".join(lines))

        return "\n\n".join(sheets) + "\n"

    def _process_csv(self, file_path: str) -> str:
        """Process CSV file to Markdown."""
        import csv
        lines = []

        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            for row_idx, row in enumerate(reader, 1):
                cells = [cell for cell in row]
                lines.append(f"| {' | '.join(cells) } |")
                if row_idx == 1:
                    lines.append("|" + "|".join(["---"] * len(cells)) + "|")

        return "\n".join(lines) + "\n"

    def _extract_images(self, file_path: str, doc_name: str):
        """Extract embedded images from Excel."""
        wb = openpyxl.load_workbook(file_path)
        for sheet in wb.worksheets:
            for img in sheet._images:
                if hasattr(img, "_data"):
                    image_data = img._data()
                    image_name = img.name or f"img_{id(img)}"
                    if not image_name.endswith((".png", ".jpg", ".jpeg", ".gif")):
                        image_name += ".png"
                    self.save_asset(doc_name, image_name, image_data)
        wb.close()